"""Collect and import the official KOSIS 2024 admin-area business workbook."""

from __future__ import annotations

import hashlib
import json
import os
import re
import shutil
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import urlopen

from openpyxl import load_workbook
from sqlalchemy import Engine, func, select
from sqlalchemy.dialects.postgresql import insert as postgresql_insert
from sqlalchemy.dialects.sqlite import insert as sqlite_insert

from localtwin_api.db_models import AdminAreaBusinessMetric, DataSource
from localtwin_api.postgres_seed import normalize_raw_path, validate_source_url
from localtwin_api.seoul_open_data import repository_root

PUBLICATION_URL = "https://kosis.kr/publication/publicationThema.do"
DOWNLOAD_URL = (
    "https://kosis.kr/upsHtml/online/downSrvcFile.do?PUBCODE=ZY&SEQ=539&FILE_NAME=0224.xlsx"
)
PERIOD = "2024"
SHEET_NAME = "읍면동 산업대분류별 총괄집계표"
WORKBOOK_FILENAME = "workbook.xlsx"
EXPECTED_COLUMN_COUNT = 59
EXPECTED_FIRST_HEADER = ("지역", "산업분류", "사업체수", "종사자수", "종사자수", "종사자수")
EXPECTED_SECOND_HEADER = ("", "", "", "계", "남", "여")
EXPECTED_INDUSTRY_CODES = ("TOTAL", *(chr(code) for code in range(ord("A"), ord("U") + 1)))
AREA_CROSSWALK = {
    "11140660": ("1144066000", "서교동"),
    "11140680": ("1144068000", "합정동"),
    "11140710": ("1144071000", "연남동"),
}
EXPECTED_SELECTED_ROWS = len(AREA_CROSSWALK) * len(EXPECTED_INDUSTRY_CODES)
INDUSTRY_PATTERN = re.compile(r"^([A-U])\.(.+)$")


class KosisBusinessCensusError(RuntimeError):
    """A sanitized workbook collection or import failure."""


@dataclass(frozen=True)
class WorkbookMetadata:
    sheet_name: str
    workbook_rows: int
    workbook_columns: int
    data_rows: int
    selected_rows: int


@dataclass(frozen=True)
class BusinessImportReport:
    snapshot_id: str
    workbook_data_rows: int
    imported_rows: int
    suppressed_rows: int
    sha256: str


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _count(value: object) -> tuple[int | None, bool]:
    normalized = _text(value).replace(",", "")
    if normalized == "X":
        return None, True
    try:
        parsed = int(normalized)
    except (TypeError, ValueError) as error:
        raise KosisBusinessCensusError("Business census count is invalid.") from error
    if parsed < 0:
        raise KosisBusinessCensusError("Business census count must not be negative.")
    return parsed, False


def _industry(value: object) -> tuple[str, str]:
    label = _text(value)
    if label == "전산업":
        return "TOTAL", label
    match = INDUSTRY_PATTERN.fullmatch(label)
    if match is None:
        raise KosisBusinessCensusError("Business census industry label changed.")
    return match.group(1), match.group(2).strip()


def parse_business_census_workbook(
    workbook_path: Path,
) -> tuple[WorkbookMetadata, list[dict[str, object]]]:
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except (OSError, ValueError) as error:
        raise KosisBusinessCensusError("Business census workbook is missing or invalid.") from error
    try:
        if workbook.sheetnames != [SHEET_NAME]:
            raise KosisBusinessCensusError("Business census sheet contract changed.")
        sheet = workbook[SHEET_NAME]
        if sheet.max_column != EXPECTED_COLUMN_COUNT:
            raise KosisBusinessCensusError("Business census column count changed.")
        first_header = tuple(_text(sheet.cell(3, column).value) for column in range(1, 7))
        second_header = tuple(_text(sheet.cell(4, column).value) for column in range(1, 7))
        if first_header != EXPECTED_FIRST_HEADER or second_header != EXPECTED_SECOND_HEADER:
            raise KosisBusinessCensusError("Business census header contract changed.")

        selected: list[dict[str, object]] = []
        seen: set[tuple[str, str]] = set()
        for row in sheet.iter_rows(min_row=5, max_col=6, values_only=True):
            region = _text(row[0])
            if "." not in region:
                continue
            source_area_code, source_area_name = region.split(".", 1)
            area = AREA_CROSSWALK.get(source_area_code)
            if area is None:
                continue
            admin_area_code, expected_name = area
            if source_area_name != expected_name:
                raise KosisBusinessCensusError("Business census admin-area name changed.")
            industry_code, industry_name = _industry(row[1])
            key = (source_area_code, industry_code)
            if key in seen:
                raise KosisBusinessCensusError(
                    "Business census contains a duplicate dimension row."
                )
            seen.add(key)
            business_count, business_suppressed = _count(row[2])
            worker_count, worker_suppressed = _count(row[3])
            male_count, male_suppressed = _count(row[4])
            female_count, female_suppressed = _count(row[5])
            if (
                worker_count is not None
                and male_count is not None
                and female_count is not None
                and worker_count != male_count + female_count
            ):
                raise KosisBusinessCensusError("Business census worker sex totals do not match.")
            selected.append(
                {
                    "admin_area_code": admin_area_code,
                    "period": PERIOD,
                    "industry_code": industry_code,
                    "admin_area_name": expected_name,
                    "source_admin_area_code": source_area_code,
                    "industry_name": industry_name,
                    "business_count": business_count,
                    "worker_count": worker_count,
                    "male_worker_count": male_count,
                    "female_worker_count": female_count,
                    "is_suppressed": any(
                        (
                            business_suppressed,
                            worker_suppressed,
                            male_suppressed,
                            female_suppressed,
                        )
                    ),
                }
            )
        expected = {
            (source_area_code, industry_code)
            for source_area_code in AREA_CROSSWALK
            for industry_code in EXPECTED_INDUSTRY_CODES
        }
        if seen != expected or len(selected) != EXPECTED_SELECTED_ROWS:
            raise KosisBusinessCensusError("Business census selected dimension set changed.")
        selected.sort(key=lambda row: (str(row["admin_area_code"]), str(row["industry_code"])))
        metadata = WorkbookMetadata(
            sheet_name=sheet.title,
            workbook_rows=sheet.max_row,
            workbook_columns=sheet.max_column,
            data_rows=sheet.max_row - 4,
            selected_rows=len(selected),
        )
        return metadata, selected
    finally:
        workbook.close()


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        while chunk := stream.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def _relative(path: Path, product_root: Path) -> str:
    try:
        return path.resolve().relative_to(product_root.resolve()).as_posix()
    except ValueError as error:
        raise KosisBusinessCensusError(
            "Business census snapshot must stay inside the product directory."
        ) from error


def write_business_census_manifest(
    workbook_path: Path,
    *,
    product_root: Path | None = None,
    collected_at: datetime | None = None,
) -> Path:
    root = product_root or repository_root()
    metadata, _ = parse_business_census_workbook(workbook_path)
    collected = collected_at or datetime.now(UTC)
    manifest = {
        "provider": "KOSIS",
        "dataset": "business-census-admin-area-industry",
        "source_type": "official-publication",
        "source_url": PUBLICATION_URL,
        "period": PERIOD,
        "collected_at": collected.strftime("%Y-%m-%dT%H:%M:%SZ"),
        "row_count": metadata.data_rows,
        "selected_row_count": metadata.selected_rows,
        "sha256": _sha256(workbook_path),
        "raw_path": _relative(workbook_path, root),
        "workbook": {
            "filename": workbook_path.name,
            "sheet": metadata.sheet_name,
            "rows": metadata.workbook_rows,
            "columns": metadata.workbook_columns,
        },
        "download_identifiers": {
            "publication_code": "ZY",
            "sequence": "539",
            "file_name": "0224.xlsx",
        },
        "area_crosswalk": AREA_CROSSWALK,
    }
    manifest_path = workbook_path.parent / "manifest.json"
    temporary = manifest_path.with_suffix(".json.tmp")
    manifest_text = json.dumps(manifest, ensure_ascii=False, indent=2) + "\n"
    temporary.write_text(manifest_text, encoding="utf-8")
    os.replace(temporary, manifest_path)
    return manifest_path


def collect_business_census_snapshot(
    output_root: Path,
    *,
    product_root: Path | None = None,
    collected_at: datetime | None = None,
    timeout_seconds: float = 120.0,
) -> Path:
    root = product_root or repository_root()
    collected = collected_at or datetime.now(UTC)
    run_dir = output_root / collected.strftime("%Y%m%dT%H%M%SZ")
    run_dir.mkdir(parents=True, exist_ok=False)
    workbook_path = run_dir / WORKBOOK_FILENAME
    temporary = workbook_path.with_suffix(".xlsx.tmp")
    try:
        response = urlopen(DOWNLOAD_URL, timeout=timeout_seconds)
        with response, temporary.open("wb") as file:
            shutil.copyfileobj(response, file)
        os.replace(temporary, workbook_path)
        write_business_census_manifest(
            workbook_path,
            product_root=root,
            collected_at=collected,
        )
    except (HTTPError, URLError, TimeoutError):
        temporary.unlink(missing_ok=True)
        raise KosisBusinessCensusError("Business census download failed.") from None
    return run_dir


def load_business_census_snapshot(
    snapshot_dir: Path,
) -> tuple[dict[str, object], WorkbookMetadata, list[dict[str, object]]]:
    try:
        manifest = json.loads((snapshot_dir / "manifest.json").read_text(encoding="utf-8"))
    except (OSError, UnicodeDecodeError, json.JSONDecodeError) as error:
        raise KosisBusinessCensusError("Business census manifest is missing or invalid.") from error
    if not isinstance(manifest, dict):
        raise KosisBusinessCensusError("Business census manifest shape is invalid.")
    expected_manifest = {
        "provider": "KOSIS",
        "dataset": "business-census-admin-area-industry",
        "source_type": "official-publication",
        "source_url": PUBLICATION_URL,
        "period": PERIOD,
        "selected_row_count": EXPECTED_SELECTED_ROWS,
    }
    if any(manifest.get(key) != value for key, value in expected_manifest.items()):
        raise KosisBusinessCensusError("Business census manifest contract changed.")
    workbook_info = manifest.get("workbook")
    if not isinstance(workbook_info, dict) or workbook_info.get("filename") != WORKBOOK_FILENAME:
        raise KosisBusinessCensusError("Business census workbook manifest is invalid.")
    try:
        validate_source_url(str(manifest["source_url"]))
        raw_path = normalize_raw_path(str(manifest["raw_path"]))
    except (KeyError, ValueError) as error:
        raise KosisBusinessCensusError("Business census provenance is unsafe.") from error
    if raw_path != manifest["raw_path"] or not raw_path.endswith(f"/{WORKBOOK_FILENAME}"):
        raise KosisBusinessCensusError("Business census raw path is invalid.")
    workbook_path = snapshot_dir / WORKBOOK_FILENAME
    if _sha256(workbook_path) != manifest.get("sha256"):
        raise KosisBusinessCensusError("Business census SHA-256 does not match the manifest.")
    metadata, rows = parse_business_census_workbook(workbook_path)
    if metadata.data_rows != manifest.get("row_count"):
        raise KosisBusinessCensusError("Business census row count does not match the manifest.")
    return manifest, metadata, rows


def _upsert(connection: Any, rows: list[dict[str, object]]) -> None:
    if connection.dialect.name == "postgresql":
        statement = postgresql_insert(AdminAreaBusinessMetric).values(rows)
    elif connection.dialect.name == "sqlite":
        statement = sqlite_insert(AdminAreaBusinessMetric).values(rows)
    else:
        raise KosisBusinessCensusError("Business census target must be PostgreSQL or test SQLite.")
    primary_keys = [column.name for column in AdminAreaBusinessMetric.__table__.primary_key.columns]
    updates = {
        column.name: getattr(statement.excluded, column.name)
        for column in AdminAreaBusinessMetric.__table__.columns
        if column.name not in primary_keys
    }
    connection.execute(statement.on_conflict_do_update(index_elements=primary_keys, set_=updates))


def import_business_census_snapshot(snapshot_dir: Path, engine: Engine) -> BusinessImportReport:
    manifest, metadata, rows = load_business_census_snapshot(snapshot_dir)
    sha256 = str(manifest["sha256"])
    snapshot_id = f"kosis-business-census-{PERIOD}-{sha256[:16]}"
    source = {
        "snapshot_id": snapshot_id,
        "provider": "KOSIS",
        "dataset": "business-census-admin-area-industry",
        "source_type": "official-publication",
        "source_url": PUBLICATION_URL,
        "collected_at": str(manifest["collected_at"]),
        "period": PERIOD,
        "row_count": metadata.data_rows,
        "sha256": sha256,
        "raw_path": str(manifest["raw_path"]),
    }
    for row in rows:
        row["source_snapshot_id"] = snapshot_id
    with engine.begin() as connection:
        if connection.dialect.name == "postgresql":
            source_insert = postgresql_insert(DataSource).values([source])
        elif connection.dialect.name == "sqlite":
            source_insert = sqlite_insert(DataSource).values([source])
        else:
            raise KosisBusinessCensusError("Business census target dialect is unsupported.")
        source_updates = {
            column.name: getattr(source_insert.excluded, column.name)
            for column in DataSource.__table__.columns
            if not column.primary_key
        }
        connection.execute(
            source_insert.on_conflict_do_update(
                index_elements=[DataSource.snapshot_id],
                set_=source_updates,
            )
        )
        _upsert(connection, rows)
        imported = int(
            connection.scalar(
                select(func.count())
                .select_from(AdminAreaBusinessMetric)
                .where(AdminAreaBusinessMetric.source_snapshot_id == snapshot_id)
            )
            or 0
        )
        if imported != EXPECTED_SELECTED_ROWS:
            raise KosisBusinessCensusError("Business census import verification failed.")
    return BusinessImportReport(
        snapshot_id=snapshot_id,
        workbook_data_rows=metadata.data_rows,
        imported_rows=imported,
        suppressed_rows=sum(bool(row["is_suppressed"]) for row in rows),
        sha256=sha256,
    )
